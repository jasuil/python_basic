# excel file read

import openpyxl as xl

excelFile = xl.load_workbook("score.xlsx")
aws = excelFile.active

print("=================")
print("name kor eng math")
print("=================")

for i in aws.rows:
    index = i[0].row
    name = i[0].value

    kor = i[1].value
    eng = i[2].value
    math = i[3].value

    total =  kor + eng + math
    avg = total / 2

    aws.cell(row=index, column=6).value = total
    aws.cell(row=index, column=7).value = avg
    print("{}, {}, {}, {}, {}, {}".format(name, kor, eng, math, total, avg))

print("*************************")

excelFile.save("scoreAppend.xlsx")